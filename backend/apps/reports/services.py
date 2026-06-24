import io
from datetime import date
from django.utils import timezone
from django.core.files.base import ContentFile
from django.conf import settings

from .models import Report


class ReportGeneratorService:
    def generate(self, report: Report) -> str:
        """Dispatch to the right generator, return file URL."""
        report.status = Report.ReportStatus.GENERATING
        report.save(update_fields=['status'])

        try:
            if report.export_format == Report.ExportFormat.PDF:
                url = self._generate_pdf(report)
            elif report.export_format == Report.ExportFormat.EXCEL:
                url = self._generate_excel(report)
            else:
                url = self._generate_json(report)

            report.status = Report.ReportStatus.COMPLETED
            report.file_url = url
            report.generated_at = timezone.now()
            report.save(update_fields=['status', 'file_url', 'generated_at'])
            return url

        except Exception as e:
            report.status = Report.ReportStatus.FAILED
            report.error_message = str(e)
            report.save(update_fields=['status', 'error_message'])
            raise

    def _gather_data(self, report: Report) -> dict:
        from apps.market.services import MarketAnalyticsService
        from apps.ai_insights.services import AIInsightsService

        data = {
            'title': report.title,
            'report_type': report.report_type,
            'generated_at': timezone.now().isoformat(),
            'user': report.user.full_name,
        }

        if report.area:
            data['area'] = report.area.name
            for ptype in ['apartment', 'villa']:
                data[f'price_trend_{ptype}'] = MarketAnalyticsService.get_price_trend(
                    report.area.slug, ptype, days=90
                )

        if report.developer:
            data['developer'] = report.developer.name
            data['developer_rankings'] = MarketAnalyticsService.get_developer_rankings()

        if report.report_type in [Report.ReportType.WEEKLY_MARKET, Report.ReportType.MONTHLY_MARKET]:
            data['inventory'] = MarketAnalyticsService.get_inventory_summary()
            data['heatmap'] = MarketAnalyticsService.get_area_heatmap()
            data['developer_rankings'] = MarketAnalyticsService.get_developer_rankings()

        # AI executive summary
        ai_svc = AIInsightsService()
        data['executive_summary'] = ai_svc.generate_executive_summary(data)

        return data

    def _generate_pdf(self, report: Report) -> str:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import cm

        data = self._gather_data(report)
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2.5*cm,
            bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        brand_blue = colors.HexColor('#1a56db')

        title_style = ParagraphStyle(
            'NileTitle',
            parent=styles['Heading1'],
            textColor=brand_blue,
            fontSize=20,
            spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            'NileHeading',
            parent=styles['Heading2'],
            textColor=brand_blue,
            fontSize=14,
            spaceBefore=16,
            spaceAfter=8,
        )

        story = []

        # Header
        story.append(Paragraph('NILE INTELLIGENCE', ParagraphStyle(
            'Brand', parent=styles['Normal'],
            textColor=brand_blue, fontSize=10, spaceAfter=4,
        )))
        story.append(Paragraph(data['title'], title_style))
        story.append(Paragraph(
            f"Generated: {data['generated_at'][:10]} | Prepared for: {data['user']}",
            styles['Normal'],
        ))
        story.append(Spacer(1, 0.5*cm))

        # Executive Summary
        story.append(Paragraph('Executive Summary', heading_style))
        story.append(Paragraph(data.get('executive_summary', ''), styles['BodyText']))
        story.append(Spacer(1, 0.5*cm))

        # Inventory Table
        if 'inventory' in data:
            story.append(Paragraph('Market Inventory', heading_style))
            inv = data['inventory']
            inv_table_data = [
                ['Metric', 'Value'],
                ['Total Active Listings', f"{inv.get('total_active', 0):,}"],
                ['New Listings (7 days)', f"{inv.get('new_last_7d', 0):,}"],
            ]
            for row in inv.get('by_type', [])[:5]:
                inv_table_data.append([row['property_type'].title(), f"{row['count']:,}"])

            t = Table(inv_table_data, colWidths=[8*cm, 8*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), brand_blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(t)

        doc.build(story)
        buffer.seek(0)

        filename = f'report_{report.id}_{date.today().isoformat()}.pdf'
        return self._upload_file(buffer.read(), filename, 'application/pdf')

    def _generate_excel(self, report: Report) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        data = self._gather_data(report)
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = 'Summary'
        header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        ws['A1'] = 'Nile Intelligence Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = data['title']
        ws['A3'] = f"Generated: {data['generated_at'][:10]}"
        ws['A5'] = 'Executive Summary'
        ws['A5'].font = Font(bold=True)
        ws['A6'] = data.get('executive_summary', '')
        ws['A6'].alignment = Alignment(wrap_text=True)
        ws.column_dimensions['A'].width = 60

        # Inventory sheet
        if 'inventory' in data:
            ws_inv = wb.create_sheet('Inventory')
            ws_inv.append(['Type', 'Count'])
            ws_inv['A1'].fill = header_fill
            ws_inv['B1'].fill = header_fill
            ws_inv['A1'].font = header_font
            ws_inv['B1'].font = header_font
            for row in data['inventory'].get('by_type', []):
                ws_inv.append([row['property_type'].title(), row['count']])

        # Price trends sheet
        for ptype in ['apartment', 'villa']:
            key = f'price_trend_{ptype}'
            if key in data and data[key]:
                ws_t = wb.create_sheet(f'Trend - {ptype.title()}')
                ws_t.append(['Date', 'Avg Price (EGP)', 'Median Price (EGP)', 'Active Listings'])
                for row_data in data[key]:
                    ws_t.append([
                        str(row_data.get('date', '')),
                        float(row_data.get('avg_price', 0) or 0),
                        float(row_data.get('median_price', 0) or 0),
                        row_data.get('active_listings', 0),
                    ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'report_{report.id}_{date.today().isoformat()}.xlsx'
        return self._upload_file(
            buffer.read(),
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _generate_json(self, report: Report) -> str:
        import json
        data = self._gather_data(report)
        content = json.dumps(data, indent=2, default=str).encode()
        filename = f'report_{report.id}_{date.today().isoformat()}.json'
        return self._upload_file(content, filename, 'application/json')

    def _upload_file(self, content: bytes, filename: str, content_type: str) -> str:
        import boto3
        from botocore.exceptions import ClientError

        s3 = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME,
        )
        key = f'reports/{filename}'
        s3.put_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            Body=content,
            ContentType=content_type,
            ACL='private',
        )
        url = s3.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': key},
            ExpiresIn=86400 * 7,
        )
        return url
